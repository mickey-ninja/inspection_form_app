# 以下のコード全部をコピーしてGitHubに貼り付けてください
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json
import os
from pathlib import Path
from PIL import Image as PILImage
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# ========== 設定 ==========
MANUAL_FILE = "93H62015_コエックス300フ_ロ_付合せ_検品包装作業.xlsx"
MASTER_FILE = "検査者マスター.xlsx"
OUTPUT_DIR = "."
PHOTO_DIR = "photos"
CONFIG_FILE = "app_config.json"

# フォルダ作成
Path(PHOTO_DIR).mkdir(parents=True, exist_ok=True)

# ========== セッション初期化 ==========
if 'inspection_data' not in st.session_state:
    st.session_state.inspection_data = {}
if 'selected_emails' not in st.session_state:
    st.session_state.selected_emails = []
if 'uploaded_photos' not in st.session_state:
    st.session_state.uploaded_photos = {}

# ========== 関数定義 ==========

def load_manual():
    """貸出検査マニュアルを読み込む"""
    try:
        wb = openpyxl.load_workbook(MANUAL_FILE)
        ws = wb.worksheets[0]
        
        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=11, max_row=45, values_only=False), 1):
            category_cell = row[0]
            description_cell = row[3]
            
            if category_cell.value or description_cell.value:
                category = category_cell.value or ""
                description = description_cell.value or ""
                
                if description.strip():
                    items.append({
                        'id': f"item_{row_idx}",
                        'category': str(category).strip(),
                        'description': str(description).strip(),
                        'row': row_idx
                    })
        
        return items
    except Exception as e:
        st.error(f"マニュアル読込エラー: {e}")
        return []

def load_masters():
    """検査者マスターを読み込む"""
    try:
        df = pd.read_excel(MASTER_FILE, sheet_name="検査者一覧")
        return df
    except Exception as e:
        st.error(f"マスター読込エラー: {e}")
        return pd.DataFrame()

def save_config(emails):
    """前回選択したメールアドレスを保存"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'selected_emails': emails}, f, ensure_ascii=False)
    except Exception as e:
        st.warning(f"設定保存エラー: {e}")

def load_config():
    """前回選択したメールアドレスを読み込む"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get('selected_emails', [])
    except:
        pass
    return []

def save_photo(uploaded_file, item_id):
    """写真を保存"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_ext = os.path.splitext(uploaded_file.name)[1]
        filename = f"{item_id}_{timestamp}{file_ext}"
        filepath = os.path.join(PHOTO_DIR, filename)
        
        with open(filepath, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        
        return filepath
    except Exception as e:
        st.error(f"写真保存エラー: {e}")
        return None

def create_excel_report(inspection_data, writer_name, reviewer_name, inspector_id, lot_no, in_no, inspection_date):
    """検査結果Excelを作成"""
    try:
        wb = openpyxl.load_workbook(MANUAL_FILE)
        ws = wb.active
        
        ws['D8'] = writer_name
        ws['P8'] = reviewer_name
        ws['D9'] = inspection_date
        ws['P9'] = inspection_date
        ws['D7'] = in_no
        ws['P7'] = lot_no
        
        result_col = 22
        for idx, (item_id, result) in enumerate(inspection_data.items()):
            row_num = 11 + idx
            if row_num < 45:
                check_value = "☑可" if result.get('pass') else "☑否"
                ws.cell(row=row_num, column=result_col).value = check_value
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"検査結果_{timestamp}.xlsx"
        wb.save(output_file)
        
        return output_file
    except Exception as e:
        st.error(f"Excel作成エラー: {e}")
        return None

# ========== UI ==========

st.set_page_config(page_title="貸出検査フォーム", layout="wide")
st.title("🔍 貸出検査フォーム")

# サイドバー設定
with st.sidebar:
    st.header("⚙️ 設定")
    
    masters = load_masters()
    if not masters.empty:
        writer_names = masters['氏名'].tolist()
        emails_list = masters['メールアドレス'].tolist()
        
        st.subheader("👤 作業者情報")
        writer_name = st.selectbox("作業者名", writer_names, key="writer")
        
        reviewer_name = st.selectbox("確認者名", writer_names, key="reviewer")
        
        st.subheader("📧 メール送信先")
        prev_emails = load_config()
        default_idx = [i for i, e in enumerate(emails_list) if e in prev_emails]
        selected_emails = st.multiselect(
            "送信先メールアドレス",
            emails_list,
            default=default_idx if default_idx else []
        )
        
        if selected_emails:
            save_config(selected_emails)
            st.session_state.selected_emails = selected_emails
    else:
        st.error("検査者マスターが見つかりません")
        writer_name = reviewer_name = None
        selected_emails = []
    
    st.subheader("📋 検査情報")
    inspector_id = st.text_input("検査ID", value=datetime.now().strftime("%Y%m%d_%H%M%S"))
    in_no = st.text_input("IN.NO", placeholder="例: IN001")
    lot_no = st.text_input("ロットNO", placeholder="例: LOT001")
    inspection_date = st.date_input("検査日", value=datetime.now())

# メインコンテンツ
manual_items = load_manual()

if not manual_items:
    st.error("検査マニュアルの読み込みに失敗しました")
else:
    st.info(f"✅ {len(manual_items)}件の検査項目を読み込みました")
    
    tabs = st.tabs(["検査入力", "確認・送信"])
    
    with tabs[0]:
        st.subheader("検査項目入力")
        
        for idx, item in enumerate(manual_items):
            with st.container():
                st.markdown(f"### No. {idx+1}: {item['category']}")
                st.write(f"📝 {item['description']}")
                
                col_check, col_photo = st.columns([2, 3])
                
                with col_check:
                    result = st.radio(
                        f"判定_{item['id']}",
                        ["可", "否"],
                        horizontal=True,
                        label_visibility="collapsed",
                        key=f"result_{item['id']}"
                    )
                    st.session_state.inspection_data[item['id']] = {
                        'description': item['description'],
                        'pass': result == "可",
                        'category': item['category']
                    }
                
                with col_photo:
                    photo = st.file_uploader(
                        f"写真アップロード_{item['id']}",
                        type=['jpg', 'jpeg', 'png'],
                        label_visibility="collapsed",
                        key=f"photo_{item['id']}"
                    )
                    
                    if photo:
                        photo_path = save_photo(photo, item['id'])
                        if photo_path:
                            st.session_state.uploaded_photos[item['id']] = photo_path
                            st.success(f"✅ 写真保存：{os.path.basename(photo_path)}")
                            img = PILImage.open(photo)
                            st.image(img, width=200)
                
                st.divider()
    
    with tabs[1]:
        st.subheader("検査結果確認")
        
        if st.session_state.inspection_data:
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                passed = sum(1 for v in st.session_state.inspection_data.values() if v.get('pass'))
                st.metric("合格項目", passed)
            
            with col2:
                failed = len(st.session_state.inspection_data) - passed
                st.metric("不合格項目", failed)
            
            with col3:
                photos = len(st.session_state.uploaded_photos)
                st.metric("写真添付数", photos)
            
            with col4:
                st.metric("検査ID", inspector_id)
            
            st.subheader("📊 検査結果一覧")
            result_df = []
            for idx, (item_id, data) in enumerate(st.session_state.inspection_data.items(), 1):
                result_df.append({
                    'No.': idx,
                    'カテゴリ': data['category'],
                    '検査項目': data['description'][:50],
                    '判定': "✅ 可" if data['pass'] else "❌ 否",
                    '写真': "📷 あり" if item_id in st.session_state.uploaded_photos else "なし"
                })
            
            result_table = pd.DataFrame(result_df)
            st.dataframe(result_table, use_container_width=True)
            
            st.subheader("💾 ファイル保存")
            
            if st.button("📊 Excel保存", use_container_width=True):
                if writer_name and reviewer_name:
                    excel_file = create_excel_report(
                        st.session_state.inspection_data,
                        writer_name, reviewer_name, inspector_id,
                        lot_no, in_no, inspection_date
                    )
                    if excel_file:
                        st.success(f"✅ Excel保存完了: {os.path.basename(excel_file)}")
                else:
                    st.error("作業者名と確認者名を選択してください")
        else:
            st.info("検査項目に回答してから確認タブをご覧ください")

st.divider()
st.caption("貸出検査フォーム v1.0 | Powered by Streamlit")
